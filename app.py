from __future__ import annotations

import sqlite3
import json
import os
import base64
import threading
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import Flask, redirect, render_template, request, send_file, url_for, make_response, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import bcrypt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from werkzeug.exceptions import HTTPException


APP_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.environ.get("DATABASE_PATH", APP_DIR / "reportabilidad.db"))
EXPORT_PATH = APP_DIR / "reporte_komatsu.xlsx"
LOGO_PATH = APP_DIR / "static" / "soldesp-logo.jpg"

TURNOS = ["Dia", "Noche"]
EQUIPOS = [
    "Sin asignacion", "Instalaciones", "Trabajos en taller de soldadura",
    "CAEX 121", "CAEX 122", "CAEX 123", "CAEX 124", "CAEX 125", "CAEX 126", "CAEX 127", "CAEX 128", "CAEX 129",
    "CAEX 130", "CAEX 131", "CAEX 132", "CAEX 133", "CAEX 134", "CAEX 135", "CAEX 136", "CAEX 137", "CAEX 138",
    "CAEX 139", "CAEX 140", "CAEX 141", "CAEX 142", "CAEX 143", "CAEX 144", "CAEX 145", "CAEX 146", "CAEX 147",
    "CAEX 148", "CAEX 149", "CAEX 150", "CAEX 151", "CAEX 152", "CAEX 153", "PALA 711", "PALA 712", "PALA 713",
    "PALA 701", "PALA 702", "PALA 703", "CARGADOR 281", "CARGADOR 282", "CARGADOR 283",
]
ACTIVIDADES = [
    "Mantenimiento preventivo estructural",
    "Reparacion estructural / soldadura",
    "Atencion de accidente / incidente",
    "Orden y aseo / housekeeping",
    "Apoyo operacional / coordinacion",
    "Traslado / logistica / espera operacional",
    "Planificacion / ART / permisos / bloqueo",
    "Capacitacion / difusion / toma conocimiento",
    "Inspeccion de seguridad / herramientas / equipos",
    "Pausas / colacion / detencion justificada",
    "Trabajos Miscelaneos",
    "Otros",
]
SUPERVISORES = [
    "Miguel Antonio Olivares Flores",
    "Hugo Orlando Guerra Puebla",
    "Gino Ronald Muñoz Gonzalez",
    "Herman Antonio Garcia Araya",
    "Rodrigo Alejandro Silva Zamora",
    "Luis Felipe Silva Vargas",
]
TRABAJADORES = [
    "Herman Patricio Contreras Castro",
    "Felipe Gomez Muñoz",
    "Juan De Dios Estay Gonzalez",
    "Rodolfo Uberlindo Jara Diaz",
    "Angel Hermindo Muñoz Ramirez",
    "Francisco Nicolas Villarroel Pulgar",
    "Joseph Vargas Delgados",
    "Eliseo Guerra Albornoz",
    "Dario Esteban Abarca Nilo",
    "Carlos Alberto Cano Meza",
    "Kevin Leandro Gonzalez Herrera",
    "Luis Osvaldo Rivera Hevia",
    "Rodrigo Alejandro Silva Zamora",
    "Fernando Ignacio Silva Gonzalez",
    "Cesar Larondo Lopez",
    "Jorge Ramirez Arancibia",
    "Daniel Peña Castillo",
    "Dilan Felipe Cavieres Rivera",
    "Yeffer Tahison Contreras Herrera",
    "Brayan Estay Perez",
    "Jose Manuel Figueroa Jara",
    "Pablo Andres Vega Gac",
    "Angelo Madariaga Arancibia",
    "Ivan Morales Morales",
    "Magdiel Gonzalo Astudillo Carvajal",
    "Patricio Espinoza Ibaceta",
    "Ricardo Antonio Flores Espinoza",
    "Alejandro Ponce Rios",
    "Jose Guillermo Saavedra Zamora",
    "Luis Felipe Silva Vargas",
    "Luciano Villegas Hevia",
    "Nicolas Antonio Lizana Olivares",
    "Felix Jamett Rivadeneira",
]
ESTADOS = ["En proceso", "Ejecutado", "Parcial", "Stand-by", "No ejecutado"]
CAUSALES = [
    "C01 Sin desviacion",
    "C02 Equipo no liberado",
    "C03 Liberacion tardia",
    "C04 Cambio prioridad Komatsu",
    "C05 Interferencia operacional",
    "C06 Falta de frente disponible",
    "C07 Falta dotacion por peak",
    "C08 Espera instruccion mandante",
    "C09 Clima adverso",
]

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "reportabilidad-local-dev")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login_page"
login_manager.login_message = "Debes iniciar sesion para acceder."

ASISTENCIA_REQUIRED_MSG = "Debe enviar asistencia para continuar"
ADMIN_REQUIRED_MSG = "Acceso restringido a administradores."
LIMITE_HH_PERSONA = 11.0


class Usuario(UserMixin):
    def __init__(self, id, username, rol, supervisor_nombre, activo):
        self.id = id
        self.username = username
        self.rol = rol
        self.supervisor_nombre = supervisor_nombre
        self.activo = activo

    def is_active(self):
        return bool(self.activo)


@login_manager.user_loader
def load_user(user_id):
    with db() as conn:
        row = conn.execute("SELECT * FROM usuarios WHERE id = ?", (int(user_id),)).fetchone()
        if row and row["activo"]:
            return Usuario(row["id"], row["username"], row["rol"], row["supervisor_nombre"], row["activo"])
    return None


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != "admin":
            return redirect(url_for("index", error=ADMIN_REQUIRED_MSG))
        return f(*args, **kwargs)
    return decorated


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def scalar(conn, query, params=()):
    row = conn.execute(query, params).fetchone()
    return row[0] or 0 if row else 0


def flash_text(value, limit=260):
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())[:limit]


def latest_sent_attendance(conn):
    return conn.execute(
        "SELECT fecha, turno, supervisor FROM asistencia WHERE estado = 'Presente' AND enviado_en IS NOT NULL ORDER BY enviado_en DESC, id DESC LIMIT 1"
    ).fetchone()


def asistencia_sent_count(conn, fecha, turno, supervisor=""):
    params = [fecha, turno]
    sf = ""
    if supervisor:
        sf = "AND supervisor = ?"
        params.append(supervisor)
    return int(scalar(conn, f"SELECT COUNT(*) FROM asistencia WHERE fecha = ? AND turno = ? AND estado = 'Presente' AND enviado_en IS NOT NULL {sf}", tuple(params)))


def requested_or_latest_context(conn):
    latest = latest_sent_attendance(conn)
    if current_user.is_authenticated and current_user.rol == "supervisor" and current_user.supervisor_nombre:
        supervisor = current_user.supervisor_nombre
    else:
        supervisor = request.values.get("supervisor") or (latest["supervisor"] if latest else "")
    fecha = request.values.get("fecha") or (latest["fecha"] if latest else date.today().isoformat())
    turno = request.values.get("turno") or (latest["turno"] if latest else "Dia")
    return {"fecha": fecha, "turno": turno, "supervisor": supervisor}


def require_asistencia_redirect(conn, endpoint="asistencia"):
    ctx = requested_or_latest_context(conn)
    if asistencia_sent_count(conn, ctx["fecha"], ctx["turno"], ctx["supervisor"]):
        return None
    return redirect(url_for(endpoint, turno=ctx["turno"], error=ASISTENCIA_REQUIRED_MSG))


def parse_hours(start, end):
    if not start or not end:
        return 0.0
    ini = datetime.strptime(start, "%H:%M")
    fin = datetime.strptime(end, "%H:%M")
    if fin < ini:
        fin += timedelta(days=1)
    return round((fin - ini).total_seconds() / 3600, 2)


def lunch_discount_hours(start, end):
    raw = parse_hours(start, end)
    if raw <= 1:
        return 0.0
    ini = datetime.strptime(start, "%H:%M")
    fin = datetime.strptime(end, "%H:%M")
    if fin < ini:
        fin += timedelta(days=1)
    return 1.0 if ini < ini.replace(hour=16, minute=0) and fin > ini.replace(hour=12, minute=0) else 0.0


def effective_hours(start, end):
    return max(0.0, round(parse_hours(start, end) - lunch_discount_hours(start, end), 2))


def classify_hh(estado, causal, hh_total, hh_directa_in, hh_indirecta_in):
    cc = (causal or "")[:3]
    d = i = n = 0.0
    adv = ""
    if estado in {"Ejecutado", "En proceso"}:
        d = hh_total
    elif estado == "Parcial":
        if hh_directa_in is None and hh_indirecta_in is None:
            d = hh_total
            adv = "Parcial sin separacion: HH clasificadas como directas."
        else:
            d = max(hh_directa_in or 0, 0)
            i = max(hh_indirecta_in or 0, 0)
            r = round(hh_total - d - i, 2)
            if r > 0:
                d += r
                adv = "Separacion parcial incompleta: saldo clasificado como HH directas."
    elif estado == "Stand-by":
        i = hh_total if cc in {"C02","C03","C04","C05","C06","C08"} else 0.0
        n = 0.0 if cc in {"C02","C03","C04","C05","C06","C08"} else hh_total
    elif estado == "No ejecutado":
        n = hh_total
    bt = "Capacidad insuficiente" if cc == "C07" else ("Condicion operacional / mandante" if cc in {"C02","C03","C04","C05","C06","C08"} else "")
    bc = {"C07": "Falta dotacion", "C02": "Falta liberacion/frente", "C03": "Falta liberacion/frente",
          "C06": "Falta liberacion/frente", "C04": "Cambio prioridad", "C05": "Interferencias", "C08": "Interferencias"}.get(cc, "")
    return {"hh_directas": round(d,2), "hh_indirectas": round(i,2), "hh_no_utilizadas": round(n,2), "advertencia": adv, "brecha_tipo": bt, "brecha_categoria": bc}


def ordenar_por_listado(nombres, listado):
    pos = {n: i for i, n in enumerate(listado)}
    return sorted(nombres, key=lambda n: (pos.get(n, len(listado)), n))


def present_count(conn, fecha, turno, cuadrilla):
    return int(scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha = ? AND turno = ? AND cuadrilla = ? AND estado = 'Presente'", (fecha, turno, cuadrilla)))


def get_options(conn, fecha=None, turno=None, supervisor=None):
    supervisores = ordenar_por_listado([r["nombre"] for r in conn.execute("SELECT nombre FROM supervisores")], SUPERVISORES)
    equipos = ordenar_por_listado([r["nombre"] for r in conn.execute("SELECT nombre FROM equipos")], EQUIPOS)
    actividades = ordenar_por_listado([r["nombre"] for r in conn.execute("SELECT nombre FROM actividades")], ACTIVIDADES)
    trabajadores = []
    if fecha and turno:
        params = [fecha, turno]
        sf = ""
        if supervisor:
            sf = "AND supervisor = ?"
            params.append(supervisor)
        trabajadores = ordenar_por_listado([r["trabajador"] for r in conn.execute(f"SELECT DISTINCT trabajador FROM asistencia WHERE fecha = ? AND turno = ? AND estado = 'Presente' AND cargo = 'Soldador' AND enviado_en IS NOT NULL {sf}", tuple(params))], TRABAJADORES)
    if not trabajadores and not (fecha and turno):
        trabajadores = ordenar_por_listado([r["nombre"] for r in conn.execute("SELECT nombre FROM trabajadores")], TRABAJADORES)
    return {
        "turnos": TURNOS, "equipos": equipos or EQUIPOS, "actividades": actividades or ACTIVIDADES,
        "estados": ESTADOS, "causales": CAUSALES, "cuadrillas": ["A","B","C","D","E"],
        "supervisores": supervisores or SUPERVISORES,
        "trabajadores": trabajadores if fecha and turno else (trabajadores or TRABAJADORES),
        "trabajadores_catalogo": ordenar_por_listado([r["nombre"] for r in conn.execute("SELECT nombre FROM trabajadores")], TRABAJADORES) or TRABAJADORES,
        "hoy": date.today().isoformat(),
    }


def init_db():
    with db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS cuadrillas (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, trabajador TEXT NOT NULL, cargo TEXT NOT NULL, especialidad TEXT NOT NULL, activo INTEGER NOT NULL DEFAULT 1);
            CREATE TABLE IF NOT EXISTS asistencia (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, turno TEXT NOT NULL, supervisor TEXT, trabajador TEXT NOT NULL, cargo TEXT NOT NULL, especialidad TEXT NOT NULL, cuadrilla TEXT NOT NULL, estado TEXT NOT NULL, hora_ingreso TEXT, hora_salida TEXT, hh_disponibles REAL NOT NULL DEFAULT 0);
            CREATE TABLE IF NOT EXISTS frentes (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, turno TEXT NOT NULL, supervisor TEXT NOT NULL, nombre_tarea TEXT, equipo TEXT NOT NULL, actividad TEXT NOT NULL, cuadrilla TEXT NOT NULL, hora_inicio TEXT NOT NULL, hora_termino TEXT NOT NULL, estado TEXT NOT NULL, causal TEXT NOT NULL, observacion TEXT, duracion_bruta REAL NOT NULL DEFAULT 0, colacion_descontada REAL NOT NULL DEFAULT 0, duracion REAL NOT NULL, personas_presentes INTEGER NOT NULL, hh_total REAL NOT NULL, hh_directas REAL NOT NULL, hh_indirectas REAL NOT NULL, hh_no_utilizadas REAL NOT NULL, advertencia TEXT, brecha_tipo TEXT, brecha_categoria TEXT, creado_en TEXT NOT NULL, art TEXT NOT NULL DEFAULT 'Si', checklist TEXT NOT NULL DEFAULT 'Si', permiso_trabajo TEXT NOT NULL DEFAULT 'Si');
            CREATE TABLE IF NOT EXISTS demanda (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT NOT NULL, turno TEXT NOT NULL, equipo TEXT NOT NULL, actividad TEXT NOT NULL, prioridad TEXT NOT NULL, estado_equipo TEXT NOT NULL, hh_requeridas REAL NOT NULL, restriccion TEXT, observacion TEXT);
            CREATE TABLE IF NOT EXISTS equipos (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE);
            CREATE TABLE IF NOT EXISTS actividades (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE);
            CREATE TABLE IF NOT EXISTS supervisores (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE);
            CREATE TABLE IF NOT EXISTS trabajadores (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE, cargo TEXT NOT NULL DEFAULT 'Soldador', especialidad TEXT NOT NULL DEFAULT 'Soldadura');
            CREATE TABLE IF NOT EXISTS usuarios (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT NOT NULL UNIQUE, password_hash TEXT NOT NULL, rol TEXT NOT NULL DEFAULT 'supervisor', supervisor_nombre TEXT, activo INTEGER NOT NULL DEFAULT 1, creado_en TEXT NOT NULL DEFAULT (datetime('now')));
        """)
        for col, defn in [("trabajadores_asignados","TEXT"),("personas_asignadas","INTEGER"),("nombre_tarea","TEXT"),("duracion_bruta","REAL NOT NULL DEFAULT 0"),("colacion_descontada","REAL NOT NULL DEFAULT 0"),("art","TEXT NOT NULL DEFAULT 'Si'"),("checklist","TEXT NOT NULL DEFAULT 'Si'"),("permiso_trabajo","TEXT NOT NULL DEFAULT 'Si'")]:
            ensure_column(conn, "frentes", col, defn)
        ensure_column(conn, "asistencia", "supervisor", "TEXT")
        ensure_column(conn, "asistencia", "enviado_en", "TEXT")
        ensure_column(conn, "asistencia", "motivo_ausencia", "TEXT")
        seed_catalogs(conn)
        if scalar(conn, "SELECT COUNT(*) FROM cuadrillas") == 0:
            seed(conn)


def ensure_column(conn, table, column, definition):
    if column not in [r["name"] for r in conn.execute(f"PRAGMA table_info({table})")]:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def seed_catalogs(conn):
    conn.executemany("INSERT OR IGNORE INTO equipos(nombre) VALUES (?)", [(x,) for x in EQUIPOS])
    conn.executemany("INSERT OR IGNORE INTO actividades(nombre) VALUES (?)", [(x,) for x in ACTIVIDADES])
    conn.execute("DELETE FROM supervisores")
    conn.executemany("INSERT OR IGNORE INTO supervisores(nombre) VALUES (?)", [(x,) for x in SUPERVISORES])
    conn.executemany("INSERT OR IGNORE INTO trabajadores(nombre) VALUES (?)", [(x,) for x in TRABAJADORES])
    if scalar(conn, "SELECT COUNT(*) FROM cuadrillas WHERE nombre = 'Soldadores'") == 0:
        conn.executemany("INSERT INTO cuadrillas(nombre, trabajador, cargo, especialidad, activo) VALUES ('Soldadores', ?, 'Soldador', 'Soldadura', 1)", [(x,) for x in TRABAJADORES])
    if scalar(conn, "SELECT COUNT(*) FROM usuarios") == 0:
        admin_user = os.environ.get("ADMIN_USER", "admin")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "Soldesp2026!")
        password_hash = bcrypt.hashpw(admin_pass.encode(), bcrypt.gensalt()).decode()
        conn.execute("INSERT INTO usuarios(username, password_hash, rol, supervisor_nombre, activo) VALUES (?, ?, 'admin', NULL, 1)", (admin_user, password_hash))


def seed(conn):
    conn.executemany("INSERT INTO cuadrillas(nombre, trabajador, cargo, especialidad, activo) VALUES (?, ?, ?, ?, 1)", [("Soldadura A","Juan Perez","Soldador","Estructural"),("Mecanica B","Ana Torres","Mecanico","Equipos")])


def frente_form_data(conn, form):
    fecha = form.get("fecha") or date.today().isoformat()
    turno = form.get("turno") or "Dia"
    cuadrilla = form.get("cuadrilla") or "Auto"
    asignados = form.getlist("trabajadores_asignados") if hasattr(form, "getlist") else []
    if not asignados and form.get("trabajadores_asignados"):
        try:
            asignados = json.loads(form.get("trabajadores_asignados") or "[]")
        except json.JSONDecodeError:
            asignados = []
    hi, ht = form.get("hora_inicio",""), form.get("hora_termino","")
    db_ = parse_hours(hi, ht)
    col = lunch_discount_hours(hi, ht)
    dur = effective_hours(hi, ht)
    asignados = asignados[:8]
    personas = len(asignados) or int(form.get("personas_asignadas") or 0) or present_count(conn, fecha, turno, cuadrilla)
    hh_total = round(dur * personas, 2)
    sd = form.get("hh_directas_split")
    si = form.get("hh_indirectas_split")
    cls = classify_hh(form.get("estado",""), form.get("causal",""), hh_total, float(sd) if sd not in (None,"") else None, float(si) if si not in (None,"") else None)
    return {"fecha":fecha,"turno":turno,"supervisor":form.get("supervisor",""),"nombre_tarea":form.get("nombre_tarea",""),"equipo":form.get("equipo",""),"actividad":form.get("actividad",""),"cuadrilla":cuadrilla,"hora_inicio":hi,"hora_termino":ht,"estado":form.get("estado",""),"causal":form.get("causal",""),"observacion":form.get("observacion",""),"duracion_bruta":db_,"colacion":col,"duracion":dur,"personas":personas,"hh_total":hh_total,"cls":cls,"asignados":asignados}


def hh_asignadas_por_persona(conn, fecha, turno, exclude_id=None):
    hh = {}
    params = [fecha, turno]
    extra = ""
    if exclude_id:
        extra = " AND id <> ?"
        params.append(exclude_id)
    for f in conn.execute(f"SELECT trabajadores_asignados, duracion FROM frentes WHERE fecha = ? AND turno = ?{extra}", tuple(params)).fetchall():
        try:
            asignados = json.loads(f["trabajadores_asignados"] or "[]")
        except Exception:
            asignados = []
        for nombre in asignados:
            hh[nombre] = round(hh.get(nombre, 0.0) + float(f["duracion"] or 0), 2)
    return hh


def hh_excedidas_guardado(conn, data, frente_id=None):
    actuales = hh_asignadas_por_persona(conn, data["fecha"], data["turno"], exclude_id=frente_id)
    return [{"nombre":n,"hh":round(actuales.get(n,0.0)+data["duracion"],2)} for n in data["asignados"] if round(actuales.get(n,0.0)+data["duracion"],2) > LIMITE_HH_PERSONA]


def insert_frente(conn, form):
    data = frente_form_data(conn, form)
    conn.execute("INSERT INTO frentes(fecha,turno,supervisor,nombre_tarea,equipo,actividad,cuadrilla,hora_inicio,hora_termino,estado,causal,observacion,duracion_bruta,colacion_descontada,duracion,personas_presentes,hh_total,hh_directas,hh_indirectas,hh_no_utilizadas,advertencia,brecha_tipo,brecha_categoria,creado_en,trabajadores_asignados,personas_asignadas,art,checklist,permiso_trabajo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (data["fecha"],data["turno"],data["supervisor"],data["nombre_tarea"],data["equipo"],data["actividad"],data["cuadrilla"],data["hora_inicio"],data["hora_termino"],data["estado"],data["causal"],data["observacion"],data["duracion_bruta"],data["colacion"],data["duracion"],data["personas"],data["hh_total"],data["cls"]["hh_directas"],data["cls"]["hh_indirectas"],data["cls"]["hh_no_utilizadas"],data["cls"]["advertencia"],data["cls"]["brecha_tipo"],data["cls"]["brecha_categoria"],datetime.now().isoformat(timespec="seconds"),json.dumps(data["asignados"],ensure_ascii=False),data["personas"],"Si","Si","Si"))


def update_frente(conn, frente_id, form):
    data = frente_form_data(conn, form)
    conn.execute("UPDATE frentes SET fecha=?,turno=?,supervisor=?,nombre_tarea=?,equipo=?,actividad=?,cuadrilla=?,hora_inicio=?,hora_termino=?,estado=?,causal=?,observacion=?,duracion_bruta=?,colacion_descontada=?,duracion=?,personas_presentes=?,hh_total=?,hh_directas=?,hh_indirectas=?,hh_no_utilizadas=?,advertencia=?,brecha_tipo=?,brecha_categoria=?,trabajadores_asignados=?,personas_asignadas=? WHERE id=?",
        (data["fecha"],data["turno"],data["supervisor"],data["nombre_tarea"],data["equipo"],data["actividad"],data["cuadrilla"],data["hora_inicio"],data["hora_termino"],data["estado"],data["causal"],data["observacion"],data["duracion_bruta"],data["colacion"],data["duracion"],data["personas"],data["hh_total"],data["cls"]["hh_directas"],data["cls"]["hh_indirectas"],data["cls"]["hh_no_utilizadas"],data["cls"]["advertencia"],data["cls"]["brecha_tipo"],data["cls"]["brecha_categoria"],json.dumps(data["asignados"],ensure_ascii=False),data["personas"],frente_id))


def validaciones_turno(conn, fecha, turno):
    presentes = {r["trabajador"]: r["cargo"] for r in conn.execute("SELECT trabajador, cargo FROM asistencia WHERE fecha = ? AND turno = ? AND estado = 'Presente' AND enviado_en IS NOT NULL", (fecha, turno))}
    hh_pp = hh_asignadas_por_persona(conn, fecha, turno)
    soldadores = [n for n, c in presentes.items() if c != "Supervisor"]
    return {
        "excedidos": [{"nombre":n,"hh":h} for n,h in hh_pp.items() if h > LIMITE_HH_PERSONA],
        "sin_actividad": [{"nombre":n,"hh":hh_pp.get(n,0.0)} for n in soldadores if hh_pp.get(n,0.0)==0.0],
        "frentes_abiertos": [dict(r) for r in conn.execute("SELECT id,nombre_tarea,equipo,estado,hora_inicio,hora_termino FROM frentes WHERE fecha=? AND turno=? AND estado='En proceso' ORDER BY hora_inicio,id",(fecha,turno))],
        "parciales": [dict(r) for r in conn.execute("SELECT id,nombre_tarea,equipo,advertencia FROM frentes WHERE fecha=? AND turno=? AND advertencia IS NOT NULL AND advertencia<>'' ORDER BY hora_inicio,id",(fecha,turno))],
        "total": 0
    }


def dashboard_data(conn, fecha, turno):
    presentes = scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente'", (fecha,turno))
    disponibles = scalar(conn, "SELECT SUM(hh_disponibles) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente'", (fecha,turno))
    directas = scalar(conn, "SELECT SUM(hh_directas) FROM frentes WHERE fecha=? AND turno=?", (fecha,turno))
    indirectas = scalar(conn, "SELECT SUM(hh_indirectas+hh_no_utilizadas) FROM frentes WHERE fecha=? AND turno=?", (fecha,turno))
    requeridas = scalar(conn, "SELECT SUM(hh_requeridas) FROM demanda WHERE fecha=? AND turno=?", (fecha,turno))
    asignadas = scalar(conn, "SELECT SUM(hh_total) FROM frentes WHERE fecha=? AND turno=?", (fecha,turno))
    equipos = scalar(conn, "SELECT COUNT(DISTINCT equipo) FROM frentes WHERE fecha=? AND turno=? AND estado IN ('Ejecutado','Parcial')", (fecha,turno))
    no_lib = scalar(conn, "SELECT COUNT(*) FROM demanda WHERE fecha=? AND turno=? AND estado_equipo IN ('No liberado','Suspendido')", (fecha,turno))
    brecha = round(requeridas - asignadas, 2)
    cats = {k: scalar(conn, f"SELECT SUM(hh_no_utilizadas+hh_indirectas) FROM frentes WHERE fecha=? AND turno=? AND brecha_categoria=?", (fecha,turno,k)) for k in ["Falta dotacion","Falta liberacion/frente","Cambio prioridad","Interferencias"]}
    sem = "Rojo" if requeridas > disponibles else ("Gris" if asignadas < disponibles*0.55 and cats["Falta liberacion/frente"] > 0 else ("Amarillo" if requeridas >= disponibles*0.9 else "Verde"))
    causales = conn.execute("SELECT causal, COUNT(*) total FROM frentes WHERE fecha=? AND turno=? GROUP BY causal ORDER BY total DESC", (fecha,turno)).fetchall()
    return {"presentes":presentes,"disponibles":round(disponibles,2),"directas":round(directas,2),"indirectas":round(indirectas,2),"requeridas":round(requeridas,2),"asignadas":round(asignadas,2),"brecha":brecha,"equipos":equipos,"no_liberados":no_lib,"semaforo":sem,"categorias":cats,"causales":causales}


@app.context_processor
def inject_today():
    try:
        with db() as conn:
            ctx = requested_or_latest_context(conn)
            nav_ready = asistencia_sent_count(conn, ctx["fecha"], ctx["turno"], ctx["supervisor"]) > 0
    except Exception:
        ctx = {"fecha": date.today().isoformat(), "turno": "Dia", "supervisor": ""}
        nav_ready = False
    return {"today": date.today().isoformat(), "nav_ctx": ctx, "nav_ready": nav_ready, "asistencia_required_msg": ASISTENCIA_REQUIRED_MSG}


@app.route("/service-worker.js")
def service_worker():
    response = make_response(send_file(APP_DIR / "static" / "sw.js"))
    response.headers["Content-Type"] = "application/javascript"
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/offline")
def offline():
    return render_template("offline.html")


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").encode()
        with db() as conn:
            row = conn.execute("SELECT * FROM usuarios WHERE username = ? AND activo = 1", (username,)).fetchone()
        if row and bcrypt.checkpw(password, row["password_hash"].encode()):
            login_user(Usuario(row["id"], row["username"], row["rol"], row["supervisor_nombre"], row["activo"]))
            return redirect(request.args.get("next") or url_for("index"))
        error = "Usuario o contraseña incorrectos."
    return render_template("login.html", error=error)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login_page"))


@app.route("/admin")
@login_required
@admin_required
def admin_panel():
    with db() as conn:
        usuarios = conn.execute("SELECT * FROM usuarios ORDER BY rol, username").fetchall()
    return render_template("admin.html", usuarios=usuarios, supervisores=SUPERVISORES, msg=request.args.get("msg"), error=request.args.get("error"))


@app.route("/admin/crear_usuario", methods=["POST"])
@login_required
@admin_required
def admin_crear_usuario():
    username = request.form.get("username","").strip()
    password = request.form.get("password","").strip()
    rol = request.form.get("rol","supervisor")
    supervisor_nombre = request.form.get("supervisor_nombre") or None
    if not username or not password or len(password) < 6:
        return redirect(url_for("admin_panel", error="Usuario y contraseña requeridos (min 6 caracteres)."))
    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    try:
        with db() as conn:
            conn.execute("INSERT INTO usuarios(username, password_hash, rol, supervisor_nombre, activo) VALUES (?,?,?,?,1)", (username, password_hash, rol, supervisor_nombre))
        return redirect(url_for("admin_panel", msg=f"Usuario '{username}' creado correctamente."))
    except Exception:
        return redirect(url_for("admin_panel", error=f"El usuario '{username}' ya existe."))


@app.route("/admin/toggle_user/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_toggle_user(uid):
    with db() as conn:
        row = conn.execute("SELECT activo, username FROM usuarios WHERE id=?", (uid,)).fetchone()
        if row and row["username"] != current_user.username:
            conn.execute("UPDATE usuarios SET activo=? WHERE id=?", (0 if row["activo"] else 1, uid))
    return redirect(url_for("admin_panel", msg="Estado actualizado."))


@app.route("/admin/reset_password/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_reset_password(uid):
    nueva = request.form.get("nueva_password", "Soldesp2026")
    password_hash = bcrypt.hashpw(nueva.encode(), bcrypt.gensalt()).decode()
    with db() as conn:
        conn.execute("UPDATE usuarios SET password_hash=? WHERE id=?", (password_hash, uid))
    return redirect(url_for("admin_panel", msg=f"Contraseña reseteada a: {nueva}"))


@app.route("/")
@login_required
def index():
    with db() as conn:
        fecha = request.args.get("fecha") or date.today().isoformat()
        turno = request.args.get("turno") or "Dia"
        supervisor = request.args.get("supervisor") or ""
        if current_user.rol == "supervisor" and current_user.supervisor_nombre:
            supervisor = current_user.supervisor_nombre
        asistencia_enviada = False
        if request.args.get("fecha") and request.args.get("turno"):
            params = [fecha, turno]
            sf = ""
            if supervisor:
                sf = "AND supervisor = ?"
                params.append(supervisor)
            asistencia_enviada = scalar(conn, f"SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND enviado_en IS NOT NULL {sf}", tuple(params)) > 0
        asistentes = scalar(conn, f"SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND enviado_en IS NOT NULL {'AND supervisor=?' if supervisor else ''}", (fecha,turno,supervisor) if supervisor else (fecha,turno)) if asistencia_enviada else 0
        actividades = scalar(conn, "SELECT COUNT(*) FROM frentes WHERE fecha=? AND turno=?", (fecha,turno)) if asistencia_enviada else 0
        validaciones = validaciones_turno(conn, fecha, turno) if asistencia_enviada else {"total": 0}
        flujo = {"asistencia_enviada": asistencia_enviada, "asistentes": asistentes, "actividades": actividades, "validaciones": validaciones, "contexto": {"fecha":fecha,"turno":turno,"supervisor":supervisor}}
        response = make_response(render_template("index.html", flujo=flujo))
        response.headers["Cache-Control"] = "no-store, max-age=0"
        return response


@app.route("/asistencia", methods=["GET", "POST"])
@login_required
def asistencia():
    with db() as conn:
        if request.method == "POST":
            fecha = request.form.get("fecha") or date.today().isoformat()
            turno = request.form.get("turno") or "Dia"
            supervisor = request.form.get("supervisor") or ""
            if current_user.rol == "supervisor" and current_user.supervisor_nombre:
                supervisor = current_user.supervisor_nombre
            accion = request.form.get("accion") or "guardar"
            ingreso = request.form.get("hora_ingreso") or ("19:15" if turno=="Noche" else "07:15")
            salida = request.form.get("hora_salida") or ("07:15" if turno=="Noche" else "19:15")
            presentes = request.form.getlist("trabajadores_presentes")
            manuales_raw = request.form.get("trabajadores_manual","")
            manuales = [x.strip() for x in manuales_raw.replace(";","\n").replace(",","\n").splitlines() if x.strip()]
            presentes = list(dict.fromkeys(presentes + manuales))
            ausentes = []
            for idx in request.form.getlist("trabajadores_ausentes"):
                nombre = request.form.get(f"ausente_nombre_{idx}","").strip()
                motivo = request.form.get(f"motivo_ausencia_{idx}","").strip()
                if nombre and nombre not in presentes:
                    ausentes.append((nombre, motivo or "Permiso"))
            hh = effective_hours(ingreso, salida)
            enviado_en = datetime.now().isoformat(timespec="seconds") if accion == "enviar" else None
            if accion == "enviar":
                conn.execute("DELETE FROM asistencia WHERE fecha < ?", (fecha,))
                conn.execute("DELETE FROM frentes WHERE fecha < ?", (fecha,))
                conn.execute("DELETE FROM demanda WHERE fecha < ?", (fecha,))
            conn.execute("DELETE FROM asistencia WHERE fecha=? AND turno=? AND supervisor=?", (fecha,turno,supervisor))
            conn.executemany("INSERT INTO asistencia(fecha,turno,supervisor,trabajador,cargo,especialidad,cuadrilla,estado,hora_ingreso,hora_salida,hh_disponibles,enviado_en,motivo_ausencia) VALUES (?,?,?,?,'Soldador','Soldadura','Turno','Presente',?,?,?,?,'')", [(fecha,turno,supervisor,t,ingreso,salida,hh,enviado_en) for t in presentes])
            conn.executemany("INSERT INTO asistencia(fecha,turno,supervisor,trabajador,cargo,especialidad,cuadrilla,estado,hora_ingreso,hora_salida,hh_disponibles,enviado_en,motivo_ausencia) VALUES (?,?,?,?,'Soldador','Soldadura','Turno','Ausente','','',0,?,?)", [(fecha,turno,supervisor,nombre,enviado_en,motivo) for nombre,motivo in ausentes])
            if supervisor:
                conn.execute("INSERT INTO asistencia(fecha,turno,supervisor,trabajador,cargo,especialidad,cuadrilla,estado,hora_ingreso,hora_salida,hh_disponibles,enviado_en,motivo_ausencia) VALUES (?,?,?,?,'Supervisor','Supervision','Supervision','Presente',?,?,?,?,'')", (fecha,turno,supervisor,supervisor,ingreso,salida,hh,enviado_en))
            conn.executemany("INSERT OR IGNORE INTO trabajadores(nombre) VALUES (?)", [(x,) for x in manuales])
            if accion == "enviar":
                conn.commit()
                run_background_task("attendance email", send_asistencia_turno, fecha, turno, supervisor)
                return redirect(url_for("frentes", fecha=fecha, turno=turno, supervisor=supervisor, msg="Asistencia enviada. Correo en camino a los destinatarios."))
            return redirect(url_for("asistencia"))
        rows = conn.execute("SELECT * FROM asistencia ORDER BY fecha DESC, turno, supervisor, enviado_en DESC, trabajador").fetchall()
        ultima_asistencia = conn.execute("SELECT fecha, turno, supervisor FROM asistencia WHERE estado='Presente' AND enviado_en IS NOT NULL ORDER BY enviado_en DESC, id DESC LIMIT 1").fetchone()
        turno_sel = request.args.get("turno","Dia")
        hora_ing = "07:15" if turno_sel == "Dia" else "19:15"
        hora_sal = "19:15" if turno_sel == "Dia" else "07:15"
        supervisor_fijo = current_user.supervisor_nombre if current_user.rol == "supervisor" else None
        return render_template("asistencia.html", rows=rows, opts=get_options(conn), turno_sel=turno_sel, hora_ing=hora_ing, hora_sal=hora_sal, ultima_asistencia=ultima_asistencia, supervisor_fijo=supervisor_fijo, error=request.args.get("error"), msg=request.args.get("msg"))


@app.route("/enviar_asistencia_pendiente", methods=["POST"])
@login_required
def enviar_asistencia_pendiente():
    fecha = request.form.get("fecha") or date.today().isoformat()
    turno = request.form.get("turno") or "Dia"
    supervisor = request.form.get("supervisor") or ""
    with db() as conn:
        conn.execute("UPDATE asistencia SET enviado_en=? WHERE fecha=? AND turno=? AND supervisor=? AND estado='Presente' AND enviado_en IS NULL", (datetime.now().isoformat(timespec="seconds"),fecha,turno,supervisor))
    return redirect(url_for("frentes", fecha=fecha, turno=turno, supervisor=supervisor))


@app.route("/cuadrillas", methods=["GET", "POST"])
@login_required
@admin_required
def cuadrillas():
    with db() as conn:
        if request.method == "POST":
            conn.execute("INSERT INTO cuadrillas(nombre,trabajador,cargo,especialidad,activo) VALUES (?,?,?,?,?)", (request.form.get("nombre"),request.form.get("trabajador"),request.form.get("cargo"),request.form.get("especialidad"),1 if request.form.get("activo") else 0))
            return redirect(url_for("cuadrillas"))
        return render_template("cuadrillas.html", rows=conn.execute("SELECT * FROM cuadrillas ORDER BY nombre, trabajador").fetchall())


@app.route("/frentes", methods=["GET", "POST"])
@login_required
def frentes():
    with db() as conn:
        gate = require_asistencia_redirect(conn)
        if gate:
            return gate
        if request.method == "POST":
            frente_id = int(request.form.get("frente_id") or 0) or None
            data = frente_form_data(conn, request.form)
            excedidos = hh_excedidas_guardado(conn, data, frente_id)
            if excedidos:
                detalle = "; ".join([f"{x['nombre']} quedaria con {x['hh']} HH" for x in excedidos])
                return redirect(url_for("frentes", fecha=data["fecha"], turno=data["turno"], supervisor=data["supervisor"], edit_id=frente_id or "", error=f"No se guardo: {detalle}."))
            if frente_id:
                update_frente(conn, frente_id, request.form)
            else:
                insert_frente(conn, request.form)
            return redirect(url_for("frentes", fecha=data["fecha"], turno=data["turno"], supervisor=data["supervisor"]))
        ultima = conn.execute("SELECT fecha, turno, supervisor FROM asistencia WHERE estado='Presente' AND enviado_en IS NOT NULL ORDER BY enviado_en DESC, id DESC LIMIT 1").fetchone()
        fecha = request.args.get("fecha") or (ultima["fecha"] if ultima else date.today().isoformat())
        turno = request.args.get("turno") or (ultima["turno"] if ultima else "Dia")
        supervisor = request.args.get("supervisor") or (ultima["supervisor"] if ultima else "")
        if current_user.rol == "supervisor" and current_user.supervisor_nombre:
            supervisor = current_user.supervisor_nombre
        rows = conn.execute("SELECT * FROM frentes WHERE fecha=? AND turno=? ORDER BY id DESC", (fecha,turno)).fetchall()
        last = conn.execute("SELECT * FROM frentes WHERE fecha=? AND turno=? ORDER BY id DESC LIMIT 1", (fecha,turno)).fetchone()
        edit_id = int(request.args.get("edit_id") or 0)
        edit = conn.execute("SELECT * FROM frentes WHERE id=?", (edit_id,)).fetchone() if edit_id else None
        edit_asignados = []
        if edit:
            try:
                edit_asignados = json.loads(edit["trabajadores_asignados"] or "[]")
            except Exception:
                pass
        pct = scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND supervisor=? AND estado='Presente' AND cargo='Soldador' AND enviado_en IS NOT NULL", (fecha,turno,supervisor)) if supervisor else scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND cargo='Soldador' AND enviado_en IS NOT NULL", (fecha,turno))
        pend = scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND supervisor=? AND estado='Presente' AND enviado_en IS NULL", (fecha,turno,supervisor)) if supervisor else scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND enviado_en IS NULL", (fecha,turno))
        return render_template("frentes.html", rows=rows, last=last, edit=edit, edit_asignados=edit_asignados, opts=get_options(conn,fecha,turno,supervisor), fecha=fecha, turno=turno, supervisor=supervisor, error=request.args.get("error"), msg=request.args.get("msg"), hh_personas=hh_asignadas_por_persona(conn,fecha,turno,exclude_id=edit_id if edit else None), present_count_turno=pct, pending_count_turno=pend)


@app.route("/demanda", methods=["GET", "POST"])
@login_required
@admin_required
def demanda():
    with db() as conn:
        if request.method == "POST":
            conn.execute("INSERT INTO demanda(fecha,turno,equipo,actividad,prioridad,estado_equipo,hh_requeridas,restriccion,observacion) VALUES (?,?,?,?,?,?,?,?,?)", (request.form.get("fecha"),request.form.get("turno"),request.form.get("equipo"),request.form.get("actividad"),request.form.get("prioridad"),request.form.get("estado_equipo"),float(request.form.get("hh_requeridas") or 0),request.form.get("restriccion"),request.form.get("observacion")))
            return redirect(url_for("demanda"))
        return render_template("demanda.html", rows=conn.execute("SELECT * FROM demanda ORDER BY fecha DESC, turno, equipo").fetchall(), opts=get_options(conn))


@app.route("/dashboard")
@login_required
def dashboard():
    fecha = request.args.get("fecha", date.today().isoformat())
    turno = request.args.get("turno", "Dia")
    with db() as conn:
        gate = require_asistencia_redirect(conn)
        if gate:
            return gate
        asistentes = conn.execute("SELECT trabajador, cargo FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND enviado_en IS NOT NULL ORDER BY CASE WHEN cargo='Supervisor' THEN 0 ELSE 1 END, trabajador", (fecha,turno)).fetchall()
        return render_template("dashboard.html", data=dashboard_data(conn,fecha,turno), asistentes=asistentes, fecha=fecha, turno=turno, opts=get_options(conn))


@app.route("/cierre")
@login_required
def cierre():
    fecha = request.args.get("fecha", date.today().isoformat())
    turno = request.args.get("turno", "Dia")
    with db() as conn:
        gate = require_asistencia_redirect(conn)
        if gate:
            return gate
        rows = conn.execute("SELECT * FROM frentes WHERE fecha=? AND turno=? ORDER BY hora_inicio, id", (fecha,turno)).fetchall()
        return render_template("cierre.html", rows=rows, data=dashboard_data(conn,fecha,turno), validaciones=validaciones_turno(conn,fecha,turno), fecha=fecha, turno=turno, opts=get_options(conn), error=request.args.get("error"), msg=request.args.get("msg"))


# ---------------------------------------------------------------------------
# Correo via Microsoft Graph API
# ---------------------------------------------------------------------------

def parse_mail_list(value):
    if not value:
        return []
    return [x.strip() for x in value.replace(";",",").replace("\n",",").split(",") if x.strip()]


def mail_recipients(kind):
    prefix = "ASSISTANCE" if kind == "asistencia" else "REPORT"
    to = parse_mail_list(os.environ.get(f"{prefix}_MAIL_TO") or os.environ.get("MAIL_TO"))
    cc = parse_mail_list(os.environ.get(f"{prefix}_MAIL_CC") or os.environ.get("MAIL_CC"))
    bcc = parse_mail_list(os.environ.get(f"{prefix}_MAIL_BCC") or os.environ.get("MAIL_BCC"))
    return to, cc, bcc


def _graph_get_token() -> str:
    """Obtiene access token de Microsoft Graph via client_credentials."""
    import urllib.request, urllib.parse
    tenant_id = os.environ.get("MS_TENANT_ID")
    client_id = os.environ.get("MS_CLIENT_ID")
    client_secret = os.environ.get("MS_CLIENT_SECRET")
    if not all([tenant_id, client_id, client_secret]):
        raise RuntimeError("Faltan variables de entorno: MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET")
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
    }).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode())["access_token"]


def send_excel_email(kind: str, subject: str, body: str, attachment_path: Path, attachment_name: str):
    """Envia correo con adjunto via Microsoft Graph API (no requiere SMTP)."""
    import urllib.request
    sender = os.environ.get("MAIL_FROM", "soldespandina@soldesp.com")
    to, cc, bcc = mail_recipients(kind)
    if not to:
        raise RuntimeError(f"Falta configurar: {'ASSISTANCE' if kind == 'asistencia' else 'REPORT'}_MAIL_TO o MAIL_TO")
    with open(attachment_path, "rb") as fh:
        attachment_b64 = base64.b64encode(fh.read()).decode()
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "Text", "content": body},
            "toRecipients": [{"emailAddress": {"address": a}} for a in to],
            "ccRecipients": [{"emailAddress": {"address": a}} for a in cc],
            "bccRecipients": [{"emailAddress": {"address": a}} for a in bcc],
            "attachments": [{
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": attachment_name,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": attachment_b64,
            }],
        },
        "saveToSentItems": "false",
    }
    token = _graph_get_token()
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        data=data,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        status = resp.status
    app.logger.info("Graph sendMail OK status=%s to=%s", status, to)
    return to + cc + bcc


def send_asistencia_turno(fecha: str, turno: str, supervisor: str):
    with db() as conn:
        enviados = scalar(conn, "SELECT COUNT(*) FROM asistencia WHERE fecha=? AND turno=? AND enviado_en IS NOT NULL", (fecha,turno))
        if not enviados:
            raise RuntimeError("Primero debes enviar la asistencia del turno.")
        ruta, nombre_archivo = build_asistencia_excel(conn, fecha, turno)
    subject = f"Asistencia turno {turno} - {fecha}"
    body = f"Estimados,\n\nSe adjunta asistencia del turno {turno} correspondiente al {fecha}.\nSupervisor: {supervisor or 'No informado'}.\n\nSaludos,\nReportabilidad Soldesp"
    return send_excel_email("asistencia", subject, body, ruta, nombre_archivo)


def run_background_task(label: str, func, *args):
    def worker():
        try:
            func(*args)
            app.logger.info("%s completed", label)
        except Exception:
            app.logger.exception("%s failed", label)
    threading.Thread(target=worker, daemon=True).start()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALERTA_FILL = PatternFill("solid", fgColor="FF4444")
ALERTA_FONT = Font(bold=True, color="FFFFFF")
WARN_FILL   = PatternFill("solid", fgColor="FFC000")
WARN_FONT   = Font(bold=False, color="000000")


def _write_sheet(ws, rows, col_widths=None):
    if not rows:
        ws.append(["Sin datos"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append([row.get(h,"") for h in headers])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+2, 45)


def _brand_existing_header(ws, title, merge_to):
    for mr in list(ws.merged_cells.ranges):
        if str(mr).startswith("A1:"):
            ws.unmerge_cells(str(mr))
    ws["A1"] = ""
    ws.row_dimensions[1].height = 42
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 14)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 14)
    if LOGO_PATH.exists():
        logo = XLImage(str(LOGO_PATH))
        logo.width = 110
        logo.height = 38
        ws.add_image(logo, "A1")
    ws["C1"] = title
    ws["C1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells(f"C1:{merge_to}1")


def _hoja_asistencia(conn, wb, fecha, turno):
    ws = wb.create_sheet("Asistencia")
    rows = []
    for r in conn.execute("SELECT trabajador, cargo, estado, motivo_ausencia, hora_ingreso, hora_salida, hh_disponibles, enviado_en FROM asistencia WHERE fecha=? AND turno=? AND enviado_en IS NOT NULL ORDER BY CASE WHEN cargo='Supervisor' THEN 0 WHEN estado='Presente' THEN 1 ELSE 2 END, trabajador", (fecha,turno)):
        rows.append({"Fecha":fecha,"Turno":turno,"Nombre":r["trabajador"],"Cargo":r["cargo"],"Estado":r["estado"],"Motivo ausencia":r["motivo_ausencia"] or "","Hora ingreso":r["hora_ingreso"],"Hora salida":r["hora_salida"],"HH disponibles":r["hh_disponibles"],"Envio asistencia":(r["enviado_en"] or "")[:16]})
    _write_sheet(ws, rows)
    ws.insert_rows(1)
    ws["A1"] = f"ASISTENCIA DIARIA  —  {turno.upper()}  —  {fecha}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 22
    _brand_existing_header(ws, f"ASISTENCIA DIARIA  -  {turno.upper()}  -  {fecha}", "J")


def _hoja_actividades(conn, wb, fecha, turno):
    ws = wb.create_sheet("Actividades")
    rows = []
    for r in conn.execute("SELECT supervisor, equipo, actividad, hora_inicio, hora_termino, duracion, personas_presentes, trabajadores_asignados, estado, causal, hh_total, hh_directas, hh_indirectas, hh_no_utilizadas, observacion FROM frentes WHERE fecha=? AND turno=? ORDER BY hora_inicio, id", (fecha,turno)):
        try:
            asignados_str = ", ".join(json.loads(r["trabajadores_asignados"] or "[]"))
        except Exception:
            asignados_str = r["trabajadores_asignados"] or ""
        rows.append({"Equipo / Frente":r["equipo"],"Actividad":r["actividad"],"Hora inicio":r["hora_inicio"],"Hora termino":r["hora_termino"],"Duracion efectiva (hrs)":r["duracion"],"Personas":r["personas_presentes"],"Trabajadores asignados":asignados_str,"Supervisor":r["supervisor"],"Estado":r["estado"],"Causal":r["causal"],"HH total":r["hh_total"],"HH directas":r["hh_directas"],"Observacion":r["observacion"] or ""})
    _write_sheet(ws, rows)
    ws.insert_rows(1)
    ws["A1"] = f"ACTIVIDADES DEL TURNO  —  {turno.upper()}  —  {fecha}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 22
    _brand_existing_header(ws, f"ACTIVIDADES DEL TURNO  -  {turno.upper()}  -  {fecha}", "M")


def _hoja_resumen_hh(conn, wb, fecha, turno):
    ws = wb.create_sheet("Resumen HH por Persona")
    presentes = {r["trabajador"]: {"cargo":r["cargo"],"hh_disponibles":r["hh_disponibles"]} for r in conn.execute("SELECT trabajador, cargo, hh_disponibles FROM asistencia WHERE fecha=? AND turno=? AND estado='Presente' AND enviado_en IS NOT NULL", (fecha,turno))}
    hh_pp = hh_asignadas_por_persona(conn, fecha, turno)
    frentes_rows = conn.execute("SELECT id FROM frentes WHERE fecha=? AND turno=?", (fecha,turno)).fetchall()
    sups = {n: i for n,i in presentes.items() if i["cargo"] == "Supervisor"}
    sols = {n: i for n,i in presentes.items() if i["cargo"] != "Supervisor"}
    rows = []
    for nombre in sorted(sups.keys()):
        rows.append({"Rol":"Supervision","Trabajador":nombre,"Cargo":"Supervisor","HH disponibles":LIMITE_HH_PERSONA,"HH asignadas":LIMITE_HH_PERSONA})
    for nombre in ordenar_por_listado(list(sols.keys()), TRABAJADORES):
        info = sols[nombre]
        hh_asig = hh_pp.get(nombre, 0.0)
        rows.append({"Rol":"Soldador","Trabajador":nombre,"Cargo":info["cargo"],"HH disponibles":info["hh_disponibles"],"HH asignadas":hh_asig})
    _write_sheet(ws, rows)
    for i, row in enumerate(rows, start=3):
        if row["Rol"] == "Supervision":
            for cell in ws[i]:
                cell.fill = PatternFill("solid", fgColor="D6E4F0")
                cell.font = Font(bold=True, color="1F4E78")
    ws.insert_rows(1)
    ws["A1"] = f"RESUMEN HH POR PERSONA  —  {turno.upper()}  —  {fecha}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 22
    _brand_existing_header(ws, f"RESUMEN HH POR PERSONA  -  {turno.upper()}  -  {fecha}", "E")


def build_reporte_excel(conn, fecha, turno):
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_asistencia(conn, wb, fecha, turno)
    _hoja_actividades(conn, wb, fecha, turno)
    _hoja_resumen_hh(conn, wb, fecha, turno)
    nombre_archivo = f"reporte_soldesp_komatsu_{fecha}_{turno.lower()}.xlsx"
    ruta = APP_DIR / nombre_archivo
    wb.save(ruta)
    return ruta, nombre_archivo


def build_asistencia_excel(conn, fecha, turno):
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_asistencia(conn, wb, fecha, turno)
    nombre_archivo = f"asistencia_soldesp_{fecha}_{turno.lower()}.xlsx"
    ruta = APP_DIR / nombre_archivo
    wb.save(ruta)
    return ruta, nombre_archivo


@app.route("/exportar")
@login_required
def exportar():
    fecha = request.args.get("fecha", date.today().isoformat())
    turno = request.args.get("turno", "Dia")
    with db() as conn:
        gate = require_asistencia_redirect(conn)
        if gate:
            return gate
        v = validaciones_turno(conn, fecha, turno)
        if v["excedidos"]:
            return redirect(url_for("cierre", fecha=fecha, turno=turno, error="No se puede descargar el informe: hay personas sobre 11 HH."))
        ruta, nombre_archivo = build_reporte_excel(conn, fecha, turno)
    return send_file(ruta, as_attachment=True, download_name=nombre_archivo)


@app.route("/exportar_asistencia")
@login_required
def exportar_asistencia():
    fecha = request.args.get("fecha", date.today().isoformat())
    turno = request.args.get("turno", "Dia")
    with db() as conn:
        ruta, nombre_archivo = build_asistencia_excel(conn, fecha, turno)
    return send_file(ruta, as_attachment=True, download_name=nombre_archivo)


@app.route("/enviar_asistencia_email", methods=["POST"])
@login_required
def enviar_asistencia_email():
    fecha = request.form.get("fecha") or date.today().isoformat()
    turno = request.form.get("turno") or "Dia"
    supervisor = request.form.get("supervisor") or ""
    next_page = request.form.get("next") or "asistencia"
    def done_url(mk, mv):
        endpoint = "frentes" if next_page == "frentes" else "asistencia"
        params = {"turno": turno, mk: mv}
        if endpoint == "frentes":
            params.update({"fecha": fecha, "supervisor": supervisor})
        return url_for(endpoint, **params)
    run_background_task("attendance email", send_asistencia_turno, fecha, turno, supervisor)
    return redirect(done_url("msg", "Envio de asistencia iniciado en segundo plano. Revisa el correo en unos minutos."))


@app.route("/enviar_reporte_email", methods=["POST"])
@login_required
def enviar_reporte_email():
    fecha = request.form.get("fecha") or date.today().isoformat()
    turno = request.form.get("turno") or "Dia"
    try:
        with db() as conn:
            gate = require_asistencia_redirect(conn)
            if gate:
                return gate
            v = validaciones_turno(conn, fecha, turno)
            if v["excedidos"]:
                raise RuntimeError("Hay personas sobre 11 HH. Edita los registros antes de enviar.")
            if v["frentes_abiertos"]:
                raise RuntimeError("Hay actividades en proceso. Cierra los registros antes de enviar.")
            ruta, nombre_archivo = build_reporte_excel(conn, fecha, turno)
        subject = f"Informe actividades turno {turno} - {fecha}"
        body = f"Estimados,\n\nSe adjunta informe de actividades del turno {turno} correspondiente al {fecha}.\n\nSaludos,\nReportabilidad Soldesp"
        run_background_task("report email", send_excel_email, "reporte", subject, body, ruta, nombre_archivo)
        return redirect(url_for("cierre", fecha=fecha, turno=turno, msg="Envio de informe iniciado en segundo plano. Revisa el correo en unos minutos."))
    except Exception as exc:
        return redirect(url_for("cierre", fecha=fecha, turno=turno, error="No se pudo enviar informe: " + flash_text(exc)))


@app.route("/limpiar_db", methods=["GET", "POST"])
@login_required
@admin_required
def limpiar_db():
    if request.method == "POST":
        with db() as conn:
            conn.executescript("DELETE FROM asistencia; DELETE FROM frentes; DELETE FROM demanda; DELETE FROM sqlite_sequence WHERE name IN ('asistencia','frentes','demanda');")
        return redirect(url_for("index"))
    return """<!doctype html><html><head><meta charset=utf-8>
    <style>body{font-family:sans-serif;max-width:480px;margin:60px auto;text-align:center}
    .btn{padding:12px 28px;border-radius:6px;font-size:1rem;cursor:pointer;border:none}
    .danger{background:#c0392b;color:#fff}.ghost{background:#eee;color:#333;text-decoration:none;padding:12px 28px;border-radius:6px}
    </style></head><body>
    <h2>&#9888; Limpiar base de datos</h2>
    <p>Se eliminarán <strong>todas las asistencias, frentes y demanda</strong>.<br>Los catálogos se conservan.</p>
    <form method=post><button class="btn danger" type=submit>Confirmar — Borrar todo</button></form><br>
    <a class=ghost href="/">Cancelar</a></body></html>"""


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    app.logger.exception("Unhandled application error")
    return redirect(url_for("asistencia", error="Error interno: " + flash_text(exc)))


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
